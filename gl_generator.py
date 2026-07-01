#!/usr/bin/env python3
"""
Cleardose Health - Dimensioned synthetic GL generator (v2).
Hybrid health-tech: SaaS subscription + patient FFS (Telehealth, CCM) + Value-Based Care capitation.
Every line carries Service_Line, State, Payor, Product dimensions. Balanced double-entry.

Usage: python gen_dim.py --start 2025-01 --end 2026-06 --out file.xlsx
       python gen_dim.py --month 2026-07 --out july.xlsx
Toggles: --ffs-drag (patient-AR collection slowdown), --no-ffs, --mlr (VBC medical loss ratio)
"""
import argparse, random, datetime as dt
from calendar import monthrange
from dateutil.relativedelta import relativedelta
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

COMPANY="Cleardose Health, Inc."
NA="—"

# ---------- dimensions ----------
STATES=["MI","OH","WI","IN","IL","MN","PA","FL","TX","AZ"]
STATE_W={"MI":1.3,"OH":1.2,"WI":0.7,"IN":0.8,"IL":1.4,"MN":0.8,"PA":1.2,"FL":1.3,"TX":1.5,"AZ":0.9}
PAYORS=["UnitedHealthcare","Humana","Cigna","Clover","Self-Pay"]
PRODUCTS=["Commercial","Medicare","Medicaid","Uninsured"]
PAYOR_PRODUCTS={
    "UnitedHealthcare":["Commercial","Medicare","Medicaid"],
    "Humana":["Medicare","Medicaid"],
    "Cigna":["Commercial","Medicare"],
    "Clover":["Medicare"],
    "Self-Pay":["Uninsured"],
}
PP_W={("UnitedHealthcare","Commercial"):1.0,("UnitedHealthcare","Medicare"):1.2,("UnitedHealthcare","Medicaid"):0.7,
      ("Humana","Medicare"):1.3,("Humana","Medicaid"):0.5,
      ("Cigna","Commercial"):0.9,("Cigna","Medicare"):0.4,
      ("Clover","Medicare"):0.6,("Self-Pay","Uninsured"):0.35}
# products each clinical service line serves
SL_PRODUCTS={"Telehealth":["Commercial","Medicare","Medicaid","Uninsured"],
             "Chronic Care Management":["Medicare","Medicaid"],
             "Value-Based Care":["Medicare","Medicaid"]}

def combos(sl):
    out=[]
    for payor in PAYORS:
        for prod in PAYOR_PRODUCTS[payor]:
            if prod in SL_PRODUCTS[sl]:
                out.append((payor,prod))
    return out

# ---------- chart of accounts ----------
COA=[
    (1000,"Cash & Cash Equivalents","BS","Asset","D"),
    (1100,"Accounts Receivable - Subscription","BS","Asset","D"),
    (1110,"Accounts Receivable - Patient/Payor","BS","Asset","D"),
    (1200,"Prepaid Expenses","BS","Asset","D"),
    (1500,"Property & Equipment, net","BS","Asset","D"),
    (2000,"Accounts Payable","BS","Liability","C"),
    (2100,"Accrued Medical Claims (IBNR)","BS","Liability","C"),
    (2150,"Accrued Liabilities","BS","Liability","C"),
    (2200,"Deferred Revenue","BS","Liability","C"),
    (3000,"Paid-in Capital","BS","Equity","C"),
    (3900,"Accumulated Deficit","BS","Equity","C"),
    (4000,"Subscription Revenue","PL","Revenue","C"),
    (4100,"Implementation & Services Revenue","PL","Revenue","C"),
    (4200,"Patient Services Revenue (FFS)","PL","Revenue","C"),
    (4250,"Capitation / VBC Revenue","PL","Revenue","C"),
    (4900,"Contractual Adjustments & Denials","PL","Revenue","D"),
    (5000,"Hosting & Infrastructure","PL","COGS","D"),
    (5100,"Clinical Provider Costs","PL","COGS","D"),
    (5150,"Care Management Staff","PL","COGS","D"),
    (5200,"Payment Processing Fees","PL","COGS","D"),
    (5300,"Third-Party Data & APIs","PL","COGS","D"),
    (5400,"Medical Claims Expense (VBC)","PL","COGS","D"),
    (6000,"R&D Salaries & Wages","PL","Opex","D"),
    (6010,"R&D Payroll Taxes & Benefits","PL","Opex","D"),
    (6100,"R&D Software & Tools","PL","Opex","D"),
    (6500,"Sales & Marketing Salaries","PL","Opex","D"),
    (6510,"S&M Payroll Taxes & Benefits","PL","Opex","D"),
    (6600,"Marketing Programs & Advertising","PL","Opex","D"),
    (6700,"Travel & Entertainment","PL","Opex","D"),
    (7000,"G&A Salaries & Wages","PL","Opex","D"),
    (7010,"G&A Payroll Taxes & Benefits","PL","Opex","D"),
    (7100,"Rent & Facilities","PL","Opex","D"),
    (7200,"Software & Subscriptions (G&A)","PL","Opex","D"),
    (7300,"Professional Fees (Legal/Audit)","PL","Opex","D"),
    (7400,"Insurance","PL","Opex","D"),
    (7500,"Office & Other G&A","PL","Opex","D"),
    (8000,"Depreciation & Amortization","PL","Opex","D"),
    (9000,"Interest Income","PL","Other Income","C"),
]
ANAME={a:n for a,n,_,_,_ in COA}; ACAT={a:c for a,_,_,c,_ in COA}; ASTMT={a:s for a,_,s,_,_ in COA}
DEPT={4000:"Revenue",4100:"Revenue",4200:"Revenue",4250:"Revenue",4900:"Revenue",
 5000:"Cost of Revenue",5100:"Cost of Revenue",5150:"Cost of Revenue",5200:"Cost of Revenue",5300:"Cost of Revenue",5400:"Cost of Revenue",
 6000:"R&D",6010:"R&D",6100:"R&D",6500:"Sales & Marketing",6510:"Sales & Marketing",6600:"Sales & Marketing",6700:"Sales & Marketing",
 7000:"G&A",7010:"G&A",7100:"G&A",7200:"G&A",7300:"G&A",7400:"G&A",7500:"G&A",8000:"G&A",9000:"G&A"}

PFX=["Mercy","Summit","Northwell","Cedar","Riverside","Beacon","Vitality","Harbor","Pioneer","Evergreen","Apex",
"Unity","Lakeshore","Capital","Compass","Meridian","Cornerstone","Bright","Sterling","Allied","Premier","Frontier",
"Cascade","Granite","Highland","Sentara"]
SFX=["Health System","Medical Group","Clinics","Care Network","Hospital","Physicians","Health Partners","Wellness",
"Diagnostics","Family Practice","Specialty Care","Health"]
random.seed(42)
def month_end(d): return dt.date(d.year,d.month,monthrange(d.year,d.month)[1])
def ps(d): return d.strftime("%Y-%m")

class Ledger:
    def __init__(self): self.rows=[]; self.jid=0
    def post(self,date,lines,memo,party="",dims=None):
        """lines: list of (acct, amt[, dimoverride]). +amt=debit, -amt=credit. dims=(SL,State,Payor,Product)."""
        s=sum(l[1] for l in lines)
        assert abs(round(s,2))<0.01,f"unbalanced {memo} {date}: {s}"
        self.jid+=1; je="JE"+str(self.jid).zfill(6)
        for l in lines:
            ac,amt=l[0],l[1]; d=l[2] if len(l)>2 and l[2] else (dims or (NA,NA,NA,NA))
            deb=round(amt,2) if amt>0 else 0.0; cre=round(-amt,2) if amt<0 else 0.0
            stmt=ASTMT[ac]
            self.rows.append(dict(JE_ID=je,Date=date,Period=ps(date),Account=ac,Account_Name=ANAME[ac],
                Statement=stmt,Category=ACAT[ac],Department=DEPT.get(ac,"G&A"),
                Service_Line=d[0],State=d[1],Payor=d[2],Product=d[3],
                Party=party,Memo=memo,Debit=deb,Credit=cre,
                PnL_Amount=round(cre-deb,2) if stmt=="PL" else 0.0,Cash_Flag=(ac==1000)))

def alloc(sl, total, rng):
    """split a monthly total across (state,payor,product) slices by weight, with noise."""
    cm=combos(sl); slices=[]; raw=[]
    for st in STATES:
        for (pay,prod) in cm:
            w=STATE_W[st]*PP_W[(pay,prod)]*rng.uniform(0.8,1.2)
            slices.append((st,pay,prod)); raw.append(w)
    tw=sum(raw)
    return [(slices[i][0],slices[i][1],slices[i][2],round(total*raw[i]/tw,2)) for i in range(len(slices))]

def simulate(start,end,ffs=True,ffs_drag=0.55,ffs_base=0.80,ffs_denial=0.07,mlr=0.85):
    rng=random.Random(42)
    L=Ledger(); od=start-dt.timedelta(days=1)
    oc,oar,opp,oppe,oap,oac,odr=18300000.0,360000.0,240000.0,420000.0,300000.0,190000.0,150000.0
    paid=26000000.0; deficit=-((oc+oar+opp+oppe)-(oap+oac+odr)-paid)
    L.post(od,[[1000,oc],[1100,oar],[1200,opp],[1500,oppe],[2000,-oap],[2150,-oac],[2200,-odr],[3000,-paid],[3900,deficit]],
           "Opening balance","Equity",("Corporate","Corporate","N/A","N/A"))
    # subscription customers (SaaS), each in a state
    customers=[]
    for i in range(70):
        st=rng.choices(STATES,weights=[STATE_W[s] for s in STATES])[0]
        customers.append({"name":rng.choice(PFX)+" "+rng.choice(SFX),"mrr":0.0,"state":st})
    w=[rng.choice([1,1,1,1,2,2,3,5,8]) for _ in customers]; tw=sum(w)
    for c,x in zip(customers,w): c["mrr"]=round(240000*x/tw,2)
    hc={"R&D":24,"Sales & Marketing":16,"G&A":11}; sal={"R&D":11800,"Sales & Marketing":11200,"G&A":12500}
    drivers=[]; cur=start; par=oar; patAR=0.0
    th_base,ccm_base,vbc_base=120000.0,60000.0,190000.0
    g=1.0
    while cur<=end:
        ld=monthrange(cur.year,cur.month)[1]; pe=dt.date(cur.year,cur.month,ld); per=ps(cur)
        # ----- SaaS subscription -----
        crate=rng.uniform(0.012,0.022); nch=max(0,round(len(customers)*crate))
        churn=rng.sample(customers,min(nch,len(customers))) if nch else []
        for c in churn: customers.remove(c)
        nnew=rng.randint(3,8); new_mrr=0.0
        for i in range(nnew):
            st=rng.choices(STATES,weights=[STATE_W[s] for s in STATES])[0]
            t=rng.choice([1,1,1,2,2,3,5]); mrr=round(rng.uniform(900,1700)*t,2); new_mrr+=mrr
            customers.append({"name":rng.choice(PFX)+" "+rng.choice(SFX),"mrr":mrr,"state":st})
        mrr_tot=round(sum(c["mrr"] for c in customers),2)
        idate=dt.date(cur.year,cur.month,1)
        for c in customers:
            if c["mrr"]>0:
                L.post(idate,[[1100,c["mrr"]],[4000,-c["mrr"]]],"Subscription invoice "+per,c["name"],
                       ("SaaS Platform",c["state"],NA,NA)); par+=c["mrr"]
        svc=round(new_mrr*rng.uniform(0.8,1.4),2)
        if svc>0:
            st=rng.choice(STATES)
            L.post(dt.date(cur.year,cur.month,8),[[1100,svc],[4100,-svc]],"Implementation services "+per,"Onboarding",
                   ("SaaS Platform",st,NA,NA)); par+=svc
        # subscription cash collection weekly
        tc=round(par*rng.uniform(0.88,0.95),2); wk=round(tc/4,2)
        for k in range(4):
            cd=dt.date(cur.year,cur.month,min(7+k*7,28)); amt=wk if k<3 else round(tc-wk*3,2)
            L.post(cd,[[1000,amt],[1100,-amt]],"Subscription cash receipts wk"+str(k+1),"Batch deposit")
            fee=round(amt*0.0125,2); L.post(cd,[[5200,fee],[1000,-fee]],"Payment processing fees","Stripe",("SaaS Platform","Corporate",NA,NA))
        par=round(par-tc,2)
        # ----- Patient FFS: Telehealth + CCM (billed gross, denials, drag collection) -----
        th_rev_by_state={s:0.0 for s in STATES}; ccm_rev_by_state={s:0.0 for s in STATES}
        gross_ffs=0.0
        for sl,base,acctcost in [("Telehealth",th_base,5100),("Chronic Care Management",ccm_base,5150)]:
            tot=base*g*rng.uniform(0.95,1.05)
            for (st,pay,prod,amt) in alloc(sl,tot,rng):
                if amt<=0: continue
                L.post(dt.date(cur.year,cur.month,3),[[1110,amt],[4200,-amt]],sl+" claims "+per,pay,(sl,st,pay,prod))
                patAR+=amt; gross_ffs+=amt
                if sl=="Telehealth": th_rev_by_state[st]+=amt
                else: ccm_rev_by_state[st]+=amt
        # denials (contra) per service line aggregate
        if ffs:
            den=round(gross_ffs*ffs_denial,2)
            L.post(dt.date(cur.year,cur.month,4),[[4900,den],[1110,-den]],"Contractual adjustments & denials","Payers",
                   ("Telehealth/CCM","Multiple","Multiple","Multiple")); patAR-=den
        # ----- VBC capitation (PMPM) + medical claims -----
        vbc_rev_by_state={s:0.0 for s in STATES}; vbc_members=0
        vtot=vbc_base*g*rng.uniform(0.95,1.05)
        for (st,pay,prod,amt) in alloc("Value-Based Care",vtot,rng):
            if amt<=0: continue
            L.post(dt.date(cur.year,cur.month,1),[[1110,amt],[4250,-amt]],"VBC capitation (PMPM) "+per,pay,
                   ("Value-Based Care",st,pay,prod)); patAR+=amt; vbc_rev_by_state[st]+=amt
            vbc_members+=int(amt/ rng.uniform(45,75))  # implied members @ PMPM
            claim=round(amt*mlr*rng.uniform(0.96,1.04),2)
            L.post(dt.date(cur.year,cur.month,20),[[5400,claim],[1000,-claim]],"Medical claims paid (VBC)","Provider network",
                   ("Value-Based Care",st,pay,prod))
        # patient AR collection with drag
        prate=0.0
        if patAR>0:
            drag=ffs_drag*(patAR/2_500_000.0); prate=ffs_base/(1+drag); coll=round(patAR*prate,2)
            for d,fr in [(12,0.5),(27,0.5)]:
                a2=round(coll*fr,2)
                if a2>0: L.post(dt.date(cur.year,cur.month,d),[[1000,a2],[1110,-a2]],"Payer remittance (patient)","Payers")
            patAR=round(patAR-coll,2)
        # ----- direct clinical costs (dimensioned by state) -----
        for st in STATES:
            if th_rev_by_state[st]>0:
                c=round(th_rev_by_state[st]*rng.uniform(0.52,0.58),2)
                L.post(pe,[[5100,c],[1000,-c]],"Telehealth provider cost","Clinician network",("Telehealth",st,NA,NA))
            if ccm_rev_by_state[st]>0:
                c=round(ccm_rev_by_state[st]*rng.uniform(0.40,0.48),2)
                L.post(pe,[[5150,c],[1000,-c]],"Care management staff","Care team",("Chronic Care Management",st,NA,NA))
        # ----- corporate COGS -----
        ncust=len(customers)
        host=round(11000+ncust*150*rng.uniform(0.95,1.08),2)
        L.post(dt.date(cur.year,cur.month,5),[[5000,host],[2000,-host]],"Hosting & infrastructure","AWS",("SaaS Platform","Corporate",NA,NA))
        da=round(6000+(gross_ffs/1000)*rng.uniform(0.9,1.1),2)
        L.post(dt.date(cur.year,cur.month,5),[[5300,da],[2000,-da]],"Clinical data & APIs","Redox/Datavant",("SaaS Platform","Corporate",NA,NA))
        # ----- payroll & opex (corporate) -----
        CORP=("Corporate","Corporate",NA,NA)
        for pd_ in [dt.date(cur.year,cur.month,15),pe]:
            for dept in hc:
                ba={"R&D":6000,"Sales & Marketing":6500,"G&A":7000}[dept]; ta={"R&D":6010,"Sales & Marketing":6510,"G&A":7010}[dept]
                gr=round(hc[dept]*sal[dept]/2*rng.uniform(0.99,1.01),2); bt=round(gr*0.24,2)
                L.post(pd_,[[ba,gr],[ta,bt],[1000,-(gr+bt)]],"Payroll "+dept,"Gusto",CORP)
        mk=round(rng.uniform(95000,160000),2)
        for v,fr in [("Google Ads",0.4),("LinkedIn",0.3),("Conferences",0.3)]:
            a2=round(mk*fr,2); L.post(dt.date(cur.year,cur.month,rng.randint(3,24)),[[6600,a2],[2000,-a2]],"Marketing programs",v,CORP)
        for ac,lo,hi,day,memo,party in [(6700,14000,38000,rng.randint(10,26),"Travel & entertainment","Brex"),
            (6100,10000,17000,4,"Engineering tools","GitHub/Datadog"),(7200,18000,28000,6,"SaaS subscriptions","Various"),
            (7400,9000,10500,2,"Insurance (D&O, cyber)","Vouch"),(7500,8000,16000,20,"Office & other G&A","Various")]:
            amt=round(rng.uniform(lo,hi),2); L.post(dt.date(cur.year,cur.month,day),[[ac,amt],[1000,-amt]],memo,party,CORP)
        L.post(dt.date(cur.year,cur.month,1),[[7100,46000],[1000,-46000]],"Office rent","Lease",CORP)
        prof=round(rng.uniform(7000,13000)+(30000 if cur.month in (3,6,9,12) else 0),2)
        L.post(dt.date(cur.year,cur.month,18),[[7300,prof],[2000,-prof]],"Legal & accounting","Cooley/KPMG",CORP)
        dep=round(oppe/48,2); L.post(pe,[[8000,dep],[1500,-dep]],"Depreciation","Non-cash",CORP)
        ic=round(rng.uniform(28000,42000),2); L.post(pe,[[1000,ic],[9000,-ic]],"Interest income","Treasury",CORP)
        appay=round(rng.uniform(220000,320000),2); L.post(dt.date(cur.year,cur.month,25),[[2000,appay],[1000,-appay]],"Vendor payments (AP)","Bill.com",CORP)
        if rng.random()<0.35:
            cx=round(rng.uniform(20000,60000),2); L.post(dt.date(cur.year,cur.month,22),[[1500,cx],[1000,-cx]],"Capital expenditure","Equipment",CORP)
        # drivers
        thc=hc["R&D"]+hc["Sales & Marketing"]+hc["G&A"]
        patrev=sum(th_rev_by_state.values())+sum(ccm_rev_by_state.values())
        vbcrev=sum(vbc_rev_by_state.values())
        dso=round(patAR/max(1,(patrev+vbcrev))*30,1)
        drivers.append(dict(Period=per,Subscription_MRR=round(mrr_tot),Subscription_ARR=round(mrr_tot*12),
            SaaS_Customers=ncust,New_Customers=nnew,Logo_Churn_Pct=round(crate,4),
            Telehealth_Rev=round(sum(th_rev_by_state.values())),CCM_Rev=round(sum(ccm_rev_by_state.values())),
            VBC_Rev=round(vbcrev),VBC_Members=vbc_members,Patient_AR=round(patAR),
            Patient_Collect_Rate=round(prate,4),Patient_DSO_Days=dso,Headcount=thc,Marketing_Spend=round(mk)))
        # growth
        g*=rng.uniform(1.04,1.07)
        if rng.random()<0.7: hc["R&D"]+=rng.randint(0,2)
        if rng.random()<0.7: hc["Sales & Marketing"]+=rng.randint(0,2)
        if rng.random()<0.4: hc["G&A"]+=rng.randint(0,1)
        cur=cur+relativedelta(months=1)
    return pd.DataFrame(L.rows),pd.DataFrame(drivers)

def classify_cf(acc,memo):
    m=memo.lower()
    if "opening balance" in m: return "Beginning Balance"
    s=set(acc)
    if 1500 in s: return "Investing"
    if s&{3000,3900} or any(k in m for k in ("raise","loan","note payable","equity round")): return "Financing"
    return "Operating"
def build_bank(gl):
    cash=gl[gl.Account==1000].copy(); others=gl[gl.Account!=1000].groupby("JE_ID").Account.apply(list).to_dict(); rows=[]
    for _,r in cash.iterrows():
        amt=round(r.Debit-r.Credit,2); acc=others.get(r.JE_ID,[]); s=set(acc)
        if s&{1100,1110}: sub="Customer / payer receipts"
        elif s&{4000,4100,4200,4250,9000}: sub="Other receipts"
        elif s&{6000,6010,6500,6510,7000,7010}: sub="Payroll"
        elif s&{5100,5150}: sub="Clinical delivery cost"
        elif 5400 in s: sub="Medical claims (VBC)"
        elif 2000 in s: sub="Vendor payments (AP)"
        elif 1500 in s: sub="Capital expenditure"
        elif s&{3000,3900}: sub="Equity / financing"
        else: sub="Operating expenses"
        rows.append(dict(Date=r.Date,Period=r.Period,JE_ID=r.JE_ID,Description=r.Memo,Counterparty=r.Party,
            Deposit=(amt if amt>0 else 0.0),Withdrawal=(-amt if amt<0 else 0.0),Amount=amt,
            CF_Category=classify_cf(acc,r.Memo),CF_Subcategory=sub))
    bk=pd.DataFrame(rows); bk["_d"]=pd.to_datetime(bk.Date); bk=bk.sort_values(["_d","JE_ID"]).reset_index(drop=True)
    bk["Running_Balance"]=bk.Amount.cumsum().round(2); bk["Date"]=bk["_d"]; bk=bk.drop(columns="_d")
    bk["Week_Starting"]=(bk.Date-pd.to_timedelta(bk.Date.dt.weekday,unit="D")).dt.date
    return bk[["Date","Week_Starting","Period","JE_ID","Description","Counterparty","Deposit","Withdrawal","Amount","Running_Balance","CF_Category","CF_Subcategory"]]
def build_budget(gl):
    # Budget mirrors actuals at the Service_Line x State grain (by period & account)
    keys=["Account","Account_Name","Category","Department","Service_Line","State"]
    pl=gl[gl.Statement=="PL"].groupby(["Period"]+keys,as_index=False).PnL_Amount.sum()
    pers=sorted(pl.Period.unique()); rows=[]; rng=random.Random(7)
    for key,gg in pl.groupby(keys):
        ac,nm,cat,dep,sl,st=key
        v=gg.set_index("Period").reindex(pers).fillna(0.0).PnL_Amount.values
        for i,per in enumerate(pers):
            plan=v[max(0,i-2):i+1].mean()*rng.uniform(0.94,1.06)
            rows.append(dict(Period=per,Account=ac,Account_Name=nm,Category=cat,Department=dep,
                Service_Line=sl,State=st,Budget_Amount=round(plan,0)))
    return pd.DataFrame(rows)

# ---------- write ----------
HF=PatternFill("solid",fgColor="1F4E79"); HFONT=Font(bold=True,color="FFFFFF",name="Arial",size=10)
MONEY='$#,##0;($#,##0);"-"'; PCT='0.0%'; DATEF='yyyy-mm-dd'; DEC1='#,##0.0'
def sty(ws,money=(),pct=(),i=(),dts=(),w=None):
    for c in range(1,ws.max_column+1):
        x=ws.cell(1,c); x.fill=HF; x.font=HFONT; x.alignment=Alignment(horizontal="center",vertical="center",wrap_text=True)
    ws.freeze_panes="A2"; ws.auto_filter.ref=ws.dimensions
    for col in money:
        for r in range(2,ws.max_row+1): ws.cell(r,col).number_format=MONEY
    for col in pct:
        for r in range(2,ws.max_row+1): ws.cell(r,col).number_format=PCT
    for col in i:
        for r in range(2,ws.max_row+1): ws.cell(r,col).number_format='#,##0'
    for col in dts:
        for r in range(2,ws.max_row+1): ws.cell(r,col).number_format=DATEF
    if w:
        for c,wd in w.items(): ws.column_dimensions[get_column_letter(c)].width=wd

def pivot_sheet(wb,name,gl,dim):
    rev=gl[gl.Category=="Revenue"].groupby([dim,"Period"]).PnL_Amount.sum().unstack(fill_value=0).round(0)
    rev["Total"]=rev.sum(axis=1); rev=rev.sort_values("Total",ascending=False)
    ws=wb.create_sheet(name); ws.cell(1,1,dim).fill=HF; ws.cell(1,1).font=HFONT
    cols=list(rev.columns)
    for j,c in enumerate(cols,2): ws.cell(1,j,str(c)); ws.cell(1,j).fill=HF; ws.cell(1,j).font=HFONT
    for i,(idx,row) in enumerate(rev.iterrows(),2):
        ws.cell(i,1,str(idx))
        for j,c in enumerate(cols,2): ws.cell(i,j,float(row[c])); ws.cell(i,j).number_format=MONEY
    tr=ws.max_row+1; ws.cell(tr,1,"Total").font=Font(bold=True,name="Arial",size=10)
    for j,c in enumerate(cols,2):
        cl=get_column_letter(j); ws.cell(tr,j,f"=SUM({cl}2:{cl}{tr-1})").number_format=MONEY; ws.cell(tr,j).font=Font(bold=True)
    ws.freeze_panes="B2"; ws.column_dimensions["A"].width=26
    for j in range(2,len(cols)+2): ws.column_dimensions[get_column_letter(j)].width=12

def write_wb(path,gl,dr,bud,bank):
    wb=load_workbook  # placeholder
    from openpyxl import Workbook
    wb=Workbook(); wb.remove(wb.active)
    def add(name,df,dates=()):
        ws=wb.create_sheet(name); ws.append(list(df.columns))
        for _,row in df.iterrows():
            ws.append([(v.to_pydatetime() if isinstance(v,pd.Timestamp) else v) for v in row.tolist()])
        return ws
    gl2=gl.copy(); gl2["Date"]=pd.to_datetime(gl2["Date"])
    add("GL_Detail",gl2); add("Bank_Activity",bank); add("Budget",bud); add("Drivers_KPIs",dr)
    add("Chart_of_Accounts",pd.DataFrame([{"Account":a,"Account_Name":n,"Statement":s,"Category":c,"Normal_Balance":nb} for a,n,s,c,nb in COA]))
    # pivots
    pivot_sheet(wb,"Rev by Service Line",gl,"Service_Line")
    pivot_sheet(wb,"Rev by State",gl,"State")
    pivot_sheet(wb,"Rev by Payor",gl,"Payor")
    pivot_sheet(wb,"Rev by Product",gl,"Product")
    # styling
    try: wb._named_styles["Normal"].font=Font(name="Arial",size=10)
    except Exception: pass
    # GL cols: JE,Date,Period,Account,Name,Stmt,Cat,Dept,SL,State,Payor,Product,Party,Memo,Debit,Credit,PnL,Cash
    sty(wb["GL_Detail"],money=(15,16,17),i=(4,),dts=(2,),
        w={1:11,2:12,3:8,4:8,5:30,6:6,7:11,8:16,9:22,10:7,11:18,12:12,13:22,14:26,15:13,16:13,17:14,18:8})
    sty(wb["Bank_Activity"],money=(7,8,9,10),dts=(1,2),w={1:12,2:13,3:8,4:11,5:28,6:22,7:13,8:13,9:13,10:15,11:13,12:24})
    sty(wb["Budget"],money=(8,),i=(2,),w={1:8,2:8,3:32,4:11,5:16,6:22,7:8,8:14})
    sty(wb["Drivers_KPIs"],money=(2,3,7,8,9,11,15),pct=(6,12),i=(4,5,10,13,14),
        w={1:8,2:15,3:15,4:14,5:14,6:13,7:14,8:12,9:12,10:13,11:13,12:14,13:14,14:12,15:15})
    sty(wb["Chart_of_Accounts"],i=(1,),w={1:8,2:34,3:10,4:14,5:14})
    # README
    rm=wb.create_sheet("README",0)
    info=[(COMPANY+" - Dimensioned GL (v2)",""),("Model","Hybrid: SaaS subscription + patient FFS (Telehealth, CCM) + Value-Based Care capitation."),
     ("",""),("NEW DIMENSIONS (every GL line)",""),
     ("Service_Line","SaaS Platform · Telehealth · Chronic Care Management · Value-Based Care · Corporate"),
     ("State","MI OH WI IN IL MN PA FL TX AZ (Corporate for overhead)"),
     ("Payor","UnitedHealthcare · Humana · Cigna · Clover · Self-Pay (revenue lines)"),
     ("Product","Commercial · Medicare · Medicaid · Uninsured"),("",""),
     ("REVENUE STREAMS",""),
     ("4000 Subscription","SaaS platform (by state; no payor/product)."),
     ("4200 Patient Services (FFS)","Telehealth + CCM claims by state/payor/product; collected via patient AR w/ drag."),
     ("4250 Capitation / VBC","Per-member-per-month risk revenue; paired with 5400 Medical Claims (MLR)."),
     ("4900 Contractual Adj.","Denials / payer adjustments (contra-revenue)."),("",""),
     ("PIVOT TABS","Rev by Service Line / State / Payor / Product (period x dimension)."),
     ("CONVENTIONS","PnL_Amount = Credit-Debit (revenue +, expense -). Every JE balances. Net rev = SUM(PnL) of Revenue category."),
     ("REFRESH","python gl_generator.py --month YYYY-MM --out file.xlsx  (toggles: --ffs-drag, --mlr, --no-ffs)"),
     ("Generated",dt.date.today().isoformat())]
    for r,(a,b) in enumerate(info,1):
        rm.cell(r,1,a); rm.cell(r,2,b)
        rm.cell(r,1).font=Font(bold=True,name="Arial",size=11 if r==1 else 10,color="1F4E79" if (r==1 or (b=="" and a.isupper())) else "000000")
        rm.cell(r,2).font=Font(name="Arial",size=10); rm.cell(r,2).alignment=Alignment(wrap_text=True,vertical="top")
    rm.column_dimensions["A"].width=26; rm.column_dimensions["B"].width=86
    order=["README","GL_Detail","Bank_Activity","Budget","Drivers_KPIs",
           "Rev by Service Line","Rev by State","Rev by Payor","Rev by Product","Chart_of_Accounts"]
    wb._sheets.sort(key=lambda s: order.index(s.title) if s.title in order else 99)
    wb.save(path)

def main():
    ap=argparse.ArgumentParser()
    ap.add_argument("--start",default="2025-01"); ap.add_argument("--end",default="2026-06"); ap.add_argument("--month",default=None)
    ap.add_argument("--out",default="Cleardose_GL_Dataset.xlsx")
    ap.add_argument("--ffs",dest="ffs",action="store_true",default=True); ap.add_argument("--no-ffs",dest="ffs",action="store_false")
    ap.add_argument("--ffs-drag",type=float,default=0.55); ap.add_argument("--ffs-denial",type=float,default=0.07)
    ap.add_argument("--mlr",type=float,default=0.85)
    a=ap.parse_args()
    if a.month: a.start=a.end=a.month
    start=dt.datetime.strptime(a.start,"%Y-%m").date().replace(day=1); end=dt.datetime.strptime(a.end,"%Y-%m").date().replace(day=1)
    gl,dr=simulate(start,end,ffs=a.ffs,ffs_drag=a.ffs_drag,ffs_denial=a.ffs_denial,mlr=a.mlr)
    bud=build_budget(gl); bank=build_bank(gl)
    write_wb(a.out,gl,dr,bud,bank)
    # validation
    assert abs(gl.Debit.sum()-gl.Credit.sum())<1,"GL not balanced"
    assert ((gl.Debit-gl.Credit).groupby(gl.JE_ID).sum().round(2).abs()<0.01).all(),"JE unbalanced"
    cash=gl[gl.Account==1000]; ec=cash.Debit.sum()-cash.Credit.sum()
    assert abs(ec-bank.Running_Balance.iloc[-1])<1,"bank!=GL cash"
    pl=gl[gl.Statement=="PL"]
    rev=pl[pl.Category=="Revenue"].PnL_Amount.sum()
    print(f"=== {COMPANY} v2 :: {a.start}..{a.end} ===")
    print(f"GL lines: {len(gl):,} | JEs: {gl.JE_ID.nunique():,} | balanced True | bank ties True")
    print(f"Ending cash: ${ec:,.0f} | Net revenue (period): ${rev:,.0f}")
    bysl=pl[pl.Category=='Revenue'].groupby('Service_Line').PnL_Amount.sum().sort_values(ascending=False)
    print("Revenue by service line:", {k:f'${v:,.0f}' for k,v in bysl.items()})
    bypp=pl[pl.Category=='Revenue'].groupby('Payor').PnL_Amount.sum().sort_values(ascending=False)
    print("Revenue by payor:", {k:f'${v:,.0f}' for k,v in bypp.items()})
    bypr=pl[pl.Category=='Revenue'].groupby('Product').PnL_Amount.sum()
    print("Revenue by product:", {k:f'${v:,.0f}' for k,v in bypr.items()})
    f=dr.iloc[-1]
    print(f"Latest: MRR ${f.Subscription_MRR:,} | Patient AR ${f.Patient_AR:,} collect {f.Patient_Collect_Rate:.0%} DSO {f.Patient_DSO_Days} | VBC members {f.VBC_Members:,}")
    print("OK")
if __name__=="__main__": main()
